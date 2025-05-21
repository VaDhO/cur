import pandas as pd
from openpyxl import load_workbook
from shutil import copyfile
import sys
import os

def save_model_to_template(model, output_path):
    import sys, os
    from openpyxl import load_workbook
    from shutil import copyfile

    def get_template_path():
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(__file__)
        return os.path.join(base_path, '1,3,4,5,6.xlsx')

    template_path = get_template_path()
    copyfile(template_path, output_path)
    wb = load_workbook(output_path)

    positions = {
        "Модель последняя": (model.tab1, 13, 3), 
        "Затраты по маслу в сх": (model.tab3, 9, 3),
        "Затраты по маслу в пр": (model.tab4, 9, 3),
        "Регулирующий блок": (model.tab5, 9, 3),
        "Управляющий блок": (model.tab6, 8, 3),
    }

    for sheet, (df, start_row, start_col) in positions.items():
        ws = wb[sheet]
        for i, row in enumerate(df.values):
            for j, value in enumerate(row):
                ws.cell(row=start_row + i, column=start_col + j, value=value)
    wb.save(output_path)
    print(f"Сохранено в файл: {output_path}")
